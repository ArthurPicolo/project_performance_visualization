import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

def find_filename(pattern, directory='.'):
    """
    Search for a file matching the given pattern in the specified directory.
    """
    for filename in os.listdir(directory):
        if re.search(pattern, filename):
            return filename
    return None

def load_and_clean_data(pl_file, project_file):
    """
    Load and clean data from P&L and Project Listing files.
    """
    # Load and clean P&L data
    pl_data = pd.read_excel(pl_file)
    pl_data = pl_data.rename(columns={
        'LABOR': 'LABOR (COGS)',
        'NON-BILLABLE': 'NON-BILLABLE EXPENSES',
        'CONT': 'CONT (MARGIN)',
        'CONT.1': 'CONT MARGIN %'
    }).drop(columns=['Unnamed: 9'], errors='ignore')

    # Remove rows with totals
    pl_data = pl_data[~pl_data['ACCOUNT MANAGER'].str.startswith('TOTAL', na=False)]

    # Load and clean project listing data
    project_data = pd.read_excel(project_file)
    project_data = project_data.rename(columns={
        'Project Name': 'PROJECT NAME',
        'Start Date': 'START DATE',
        'Finish Date': 'END DATE',
        'Project Status': 'STATUS'
    })

    # Merge datasets
    merged_data = pd.merge(pl_data, project_data, on='PROJECT NAME', how='left')
    
    # Select and reorder columns
    final_columns = [
        'ACCOUNT MANAGER', 'CUSTOMER', 'PROJECT NAME', 
        'START DATE', 'END DATE', 'STATUS', 'LINE OF BUSINESS', 
        'REVENUE', 'INVESTMENT', 'LABOR (COGS)', 
        'NON-BILLABLE EXPENSES', 'CONT (MARGIN)', 'CONT MARGIN %'
    ]
    return merged_data[final_columns]

def apply_excel_formatting(file_path):
    """
    Apply formatting to an Excel file.
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # Define formatting
    columns_to_currency = ['H', 'I', 'J', 'K', 'L']
    columns_to_date = ['D', 'E']
    columns_to_percentage = ['M']
    ws.delete_rows(2)

    # Apply formatting
    for row in range(2, ws.max_row + 1):
        for col in columns_to_currency:
            ws[f'{col}{row}'].number_format = 'R$ #,##0.00'

        for col in columns_to_date:
            ws[f'{col}{row}'].number_format = 'DD/MM/YYYY'

        for col in columns_to_percentage:
            ws[f'{col}{row}'].number_format = '0.00%'

    wb.save(file_path)

def main():
    # Find P&L file
    pl_file_pattern = r'Processed_PNL.*\.xlsx'
    pl_file = find_filename(pl_file_pattern)
    
    if not pl_file:
        print("The following file was not found: 'P L By Account Manager _PL By AM'.")
        return

    # Define project listing file
    project_file = 'Processed_ProjectListing.xlsx'
    if not os.path.exists(project_file):
        print(f"The file '{project_file}' was not found.")
        return

    # Load and clean data
    final_data = load_and_clean_data(pl_file, project_file)
    print(type(final_data))

    # Save to Excel
    output_file = 'PL_By_AM_Final.xlsx'
    final_data.to_excel(output_file, index=False)

    # Apply formatting
    apply_excel_formatting(output_file)

    print(f"Processed file saved as: {output_file}")

if __name__ == '__main__':
    main()
